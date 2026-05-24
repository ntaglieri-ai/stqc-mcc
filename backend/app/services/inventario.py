from pathlib import Path
from typing import Any, List


def parse_inventario(file_path: Path) -> List[dict]:
    """Legge il foglio 'Inventario ...' dal file .xlsm e restituisce le righe
    con stock attuale > 0 come lista di dizionari pronti per l'import."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=file_path, data_only=True, keep_vba=False)

    # Trova il foglio inventario (il primo che non è PIVOT/NOTE)
    sheet = None
    for name in wb.sheetnames:
        if name.upper() not in ("PIVOT", "NOTE"):
            sheet = wb[name]
            break
    if sheet is None:
        return []

    # Colonne attese (indice 0-based):
    # 0=TIPO  1=PROFILO  2=N°PEZZI  3=DIMENSIONI  4=QUALITA'  5=COLATA
    # 6=COMMESSA  7=PESO[KG]  8=PESO/U  9=PESO1PZ  10=PAGINA
    # 11=DATA_PREL  12=QTA_PREL  13=DDT/RESIDUO  14=DATA_ARRIVO
    # 15=QTA_ARRIVATA  16=N°PEZZI_ATTUALE

    items: List[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tipo = row[0]
        if not tipo or not isinstance(tipo, str):
            continue

        profilo = row[1]
        pezzi_attuale = row[16]

        try:
            qty = float(pezzi_attuale) if pezzi_attuale not in (None, "") else 0.0
        except (TypeError, ValueError):
            qty = 0.0

        if qty <= 0:
            continue

        def _str(v) -> str | None:
            return str(v).strip() if v not in (None, "") else None

        def _float(v) -> float | None:
            try:
                return float(v) if v not in (None, "") else None
            except (TypeError, ValueError):
                return None

        tipo_str = tipo.strip()
        profilo_str = _str(profilo) or ""
        qualita = _str(row[4])
        colata = _str(row[5])
        commessa = _str(row[6])
        dimensioni = _str(row[3])
        peso_u_kg = _float(row[8])
        peso_1_pz = _float(row[9])

        # Codice materiale univoco: TIPO-PROFILO[-QUALITA'] troncato a 100 char
        parts = [tipo_str.upper().replace(" ", "_"), profilo_str.replace(" ", "_")]
        if qualita:
            parts.append(qualita.replace(" ", "_"))
        material_code = "-".join(parts)[:100]

        # Descrizione leggibile (retrocompatibilità)
        desc_parts = [f"{tipo_str} {profilo_str}".strip()]
        if qualita:
            desc_parts.append(qualita)
        if dimensioni:
            desc_parts.append(f"L={dimensioni}mm")
        description = " | ".join(desc_parts)[:400]

        items.append({
            "material_code": material_code,
            "description": description,
            "specification": qualita,
            # campi strutturati
            "tipo": tipo_str,
            "profilo": profilo_str or None,
            "dimensioni": dimensioni,
            "qualita": qualita,
            "colata": colata,
            "commessa_reference": commessa,
            "peso_u_kg": peso_u_kg,
            "peso_1_pz": peso_1_pz,
            "quantity": qty,
        })

    return items
