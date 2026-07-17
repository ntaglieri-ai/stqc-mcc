"""Parser distinta base — gestisce due file Tekla/Advanced Steel.

Output normalizzato (indipendente dal CAD):
    {
        "part_code":       str,
        "profile":         str | None,
        "weight_kg":       float | None,
        "qty":             int,          # totale da espandere in istanze
        "length_mm":       float | None,
        "width_mm":        float | None,
        "material":        str | None,
        "assembly_parent": str | None,
    }

Ogni parte con qty > 1 viene espansa in N record separati
(instance_number 1..N), ciascuno con "qty": 1.
"""
from __future__ import annotations

import re
from collections import defaultdict, deque
from html import unescape
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, List

import xlrd
from openpyxl import load_workbook


# ── Normalizzazione profili ───────────────────────────────────────────────────

_PROFILE_ALIASES: dict[str, str] = {
    # Lamiere
    "LAMIERA10": "PL10", "P10": "PL10",
    "LAMIERA15": "PL15", "P15": "PL15",
    "LAMIERA20": "PL20", "P20": "PL20",
    "LAMIERA25": "PL25", "P25": "PL25",
    "LAMIERA30": "PL30", "P30": "PL30",
    # Forme comuni con spazio o varianti maiuscolo
    "IPE 100": "IPE100", "IPE 120": "IPE120", "IPE 140": "IPE140",
    "IPE 160": "IPE160", "IPE 180": "IPE180", "IPE 200": "IPE200",
    "IPE 220": "IPE220", "IPE 240": "IPE240", "IPE 270": "IPE270",
    "IPE 300": "IPE300", "IPE 330": "IPE330", "IPE 360": "IPE360",
    "IPE 400": "IPE400", "IPE 450": "IPE450", "IPE 500": "IPE500",
    "HEA 100": "HEA100", "HEB 100": "HEB100",
    "HEA 120": "HEA120", "HEB 120": "HEB120",
    "HEA 140": "HEA140", "HEB 140": "HEB140",
    "HEA 160": "HEA160", "HEB 160": "HEB160",
    "HEA 180": "HEA180", "HEB 180": "HEB180",
    "HEA 200": "HEA200", "HEB 200": "HEB200",
    "HEA 220": "HEA220", "HEB 220": "HEB220",
    "HEA 240": "HEA240", "HEB 240": "HEB240",
    "HEA 260": "HEA260", "HEB 260": "HEB260",
    "HEA 280": "HEA280", "HEB 280": "HEB280",
    "HEA 300": "HEA300", "HEB 300": "HEB300",
    "UPN 100": "UPN100", "UPN 120": "UPN120", "UPN 140": "UPN140",
    "UPN 160": "UPN160", "UPN 180": "UPN180", "UPN 200": "UPN200",
    "UPN 220": "UPN220", "UPN 240": "UPN240", "UPN 260": "UPN260",
    "UPN 300": "UPN300",
    "L 50X50X5": "L50X50X5", "L 60X60X6": "L60X60X6",
    "L 70X70X7": "L70X70X7", "L 80X80X8": "L80X80X8",
    "L 100X100X10": "L100X100X10",
    # Tubi quadri/rettangolari (TUBE-C o RHS)
    "TUBOC": "TUBE-C", "RHS": "TUBE-C",
}


def normalize_profile(value: str | None) -> str:
    """Normalizza un codice profilo in forma canonica per confronti DB-agnostici.

    Regole:
    - UPPER + strip degli spazi laterali
    - Elimina tutti gli spazi interni
    - Elimina i trattini
    - Risolve alias noti (lamiere P→PL, varianti spaziate IPE/HEA/HEB/UPN)
    Aggiungere alias a _PROFILE_ALIASES man mano che emergono dalle anomalie.
    """
    if not value:
        return ""
    v = value.upper().strip()
    v = re.sub(r'\s+', '', v)
    v = v.replace('-', '')
    return _PROFILE_ALIASES.get(v, v)


# ── Alias colonne ─────────────────────────────────────────────────────────────

ALIASES: dict[str, list[str]] = {
    "part_code":    ["marca/pos.", "posizione", "part", "codice", "position", "parte", "marca/pos"],
    "assembly":     ["assemb.", "assemb", "assembly", "assemblato", "asemb.", "asemb"],
    "profile":      ["profilo", "section", "profile", "descrizione", "desc"],
    "qty":          ["q.tà", "q.ta", "qty", "quantity", "quantita", "n°", "qta"],
    # weight: il file assemblaggi usa "Peso(kg) un.", il file lavorazioni usa "Peso Netto (kg) per uno"
    "weight":       ["peso(kg) un.", "peso(kg)un.", "peso kg un.", "peso", "weight",
                     "peso unit.", "peso(kg)", "peso netto (kg) per uno", "peso netto(kg) per uno"],
    # length_mm: il file assemblaggi usa "Lungh. (mm)" (con spazio e parentesi),
    #            il file lavorazioni usa "Lungh." (senza parentesi)
    "length_mm":    ["lunghezza", "lungh.", "lungh. (mm)", "lung.", "length", "lung", "l(mm)", "l mm"],
    "width_mm":     ["larghezza", "largh.", "largh. (mm)", "width", "w(mm)", "w mm"],
    "material":     ["materiale", "material", "qualità", "qualita", "quality"],
    "commessa_ref": ["commessa", "commessa_reference", "commessa_ref", "order"],
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _norm(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _is_valid_part_code(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if not s or len(s) > 30 or len(s.split()) > 2:
        return False
    return any(c.isdigit() for c in s)


class _HtmlTableRowsParser(HTMLParser):
    """Estrae celle da file Excel esportati come HTML ma salvati con estensione .xls."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None
        self._in_table = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._in_table = True
        elif self._in_table and tag == "tr":
            self._current_row = []
        elif self._in_table and tag in {"td", "th"} and self._current_row is not None:
            self._current_cell = []
        elif self._in_table and tag == "br" and self._current_cell is not None:
            self._current_cell.append(" ")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._current_cell is not None and self._current_row is not None:
            text = unescape("".join(self._current_cell))
            text = re.sub(r"\s+", " ", text).strip()
            self._current_row.append(text)
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None:
            if any(cell != "" for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None
        elif tag == "table":
            self._in_table = False

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell.append(data)


def _looks_like_html_excel(raw: bytes) -> bool:
    sample = raw[:2048].lstrip().lower()
    return sample.startswith(b"<!doctype html") or sample.startswith(b"<html") or b"<table" in sample


def _extract_html_table_rows(file_path: Path) -> list[list[Any]]:
    raw = file_path.read_bytes()
    for encoding in ("utf-8-sig", "iso-8859-1", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    parser = _HtmlTableRowsParser()
    parser.feed(text)
    return parser.rows


def _extract_rows(file_path: Path) -> list[list[Any]]:
    suffix = file_path.suffix.lower()
    raw = file_path.read_bytes()
    if _looks_like_html_excel(raw):
        return _extract_html_table_rows(file_path)
    if suffix == ".xls":
        wb = xlrd.open_workbook(file_path.as_posix(), formatting_info=False)
        sh = wb.sheet_by_index(0)
        return [sh.row_values(r) for r in range(sh.nrows)]
    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    sh = wb.active
    return [list(row) for row in sh.iter_rows(values_only=True)]


def _find_header_row(rows: list[list[Any]], key_aliases: list[str]) -> int:
    """Trova la riga che contiene almeno uno degli alias forniti."""
    for i, row in enumerate(rows[:20]):
        normalized = [_norm(c) for c in row]
        if any(alias in normalized for alias in key_aliases):
            return i
    return 0


def _detect_file_type(path: Path) -> str:
    """Rileva il tipo di file analizzando le colonne della riga header.

    Regole (in ordine di priorità):
    - Ha colonna "Assemb." E colonna "Parte"/"Marca/Pos." in colonne distinte
      → "assemblaggi"  (Lista parti assemblaggi — gerarchia assemblato→parti)
    - Ha colonna "Marca/Pos." ma NON "Assemb." separata
      → "lavorazioni"  (Lavorazioni per posizione — dettagli tecnici)
    - Altrimenti → "unknown"  (usa l'ordine passato come fallback)
    """
    try:
        rows = _extract_rows(path)
    except Exception:
        return "unknown"
    if not rows:
        return "unknown"

    all_triggers = ALIASES["assembly"] + ALIASES["part_code"] + ALIASES["qty"]
    header_idx = _find_header_row(rows, all_triggers)
    col_map = _build_col_map(rows[header_idx])

    has_assembly  = "assembly"  in col_map
    has_part_code = "part_code" in col_map
    separate_cols = (
        has_assembly and has_part_code
        and col_map["assembly"] != col_map["part_code"]
    )

    if separate_cols:
        return "assemblaggi"
    if has_part_code:          # Marca/Pos. senza Assemb. separata
        return "lavorazioni"
    if has_assembly:           # Solo colonna assembly → trattare come assemblaggi
        return "assemblaggi"
    return "unknown"


def _build_col_map(header: list[Any]) -> dict[str, int]:
    """Mappa nome-campo → indice colonna usando ALIASES."""
    col_map: dict[str, int] = {}
    for ix, cell in enumerate(header):
        h = _norm(cell)
        for field, aliases in ALIASES.items():
            if field not in col_map and h in aliases:
                col_map[field] = ix
                break
    return col_map


def _float_cell(row: list[Any], col_map: dict[str, int], key: str) -> float | None:
    if key not in col_map:
        return None
    ix = col_map[key]
    if ix >= len(row):
        return None
    val = row[ix]
    if val in (None, ""):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        try:
            cleaned = re.sub(r"[^\d.,\-]", "", str(val)).replace(",", ".")
            return float(cleaned) if cleaned else None
        except ValueError:
            return None


def _str_cell(row: list[Any], col_map: dict[str, int], key: str) -> str | None:
    if key not in col_map:
        return None
    ix = col_map[key]
    if ix >= len(row):
        return None
    val = row[ix]
    return str(val).strip() if val not in (None, "") else None


def _find_commessa(rows: list[list[Any]]) -> str | None:
    for row in rows[:12]:
        for i, cell in enumerate(row):
            if isinstance(cell, str) and "commessa" in cell.lower():
                nxt = row[i + 1] if i + 1 < len(row) else None
                if nxt not in (None, ""):
                    return str(nxt).strip()
    return None


# ── Parser singolo file (legacy + backward compat) ───────────────────────────

def _parse_single(file_path: Path) -> tuple[list[dict], set[str]]:
    """Parser generico per un singolo file.
    Restituisce (items_normalizzati, assembly_headers) dove assembly_headers è
    l'insieme dei codici che compaiono come testata-assemblato (non come parti).
    """
    rows = _extract_rows(file_path)
    if not rows:
        return [], set()

    all_part_aliases = ALIASES["part_code"] + ALIASES["assembly"]
    header_idx = _find_header_row(rows, all_part_aliases + ALIASES["qty"])
    header = rows[header_idx]
    col_map = _build_col_map(header)

    has_separate_assembly = (
        "assembly" in col_map
        and "part_code" in col_map
        and col_map["assembly"] != col_map["part_code"]
    )

    commessa_ref = _find_commessa(rows)
    current_assembly: str | None = None
    assembly_headers: set[str] = set()
    items: list[dict] = []

    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        if has_separate_assembly:
            asm_val = _str_cell(row, col_map, "assembly")
            pn_val  = _str_cell(row, col_map, "part_code")
            if _is_valid_part_code(asm_val) and not pn_val:
                current_assembly = asm_val
                assembly_headers.add(asm_val)
                continue

        raw_pn = _str_cell(row, col_map, "part_code")
        if not _is_valid_part_code(raw_pn):
            continue

        qty_raw = _float_cell(row, col_map, "qty")
        n = max(1, int(round(qty_raw))) if qty_raw and qty_raw > 0 else 1

        base: dict = {
            "part_code":       raw_pn,
            "profile":         normalize_profile(_str_cell(row, col_map, "profile")),
            "weight_kg":       _float_cell(row, col_map, "weight"),
            "length_mm":       _float_cell(row, col_map, "length_mm"),
            "width_mm":        _float_cell(row, col_map, "width_mm"),
            "material":        _str_cell(row, col_map, "material"),
            "assembly_parent": current_assembly if has_separate_assembly else _str_cell(row, col_map, "assembly"),
            "commessa_ref":    _str_cell(row, col_map, "commessa_ref") or commessa_ref,
        }

        for inst in range(1, n + 1):
            items.append({**base, "qty": 1, "instance_number": inst})

    return items, assembly_headers


# ── Parser commessa ───────────────────────────────────────────────────────────

def _parse_assembly_hierarchy(file_path: Path) -> tuple[dict[str, deque[str]], set[str], list[str]]:
    """Restituisce le assegnazioni part_code → coda assemblati.

    La quantità fisica di una parte è:
        quantità assemblato × quantità parte per assemblato

    Il file assemblaggi arricchisce la lista pezzi con il parent, ma non decide
    mai il numero totale di pezzi: la sorgente autoritativa resta il file
    "Lavorazioni per posizione".
    """
    rows = _extract_rows(file_path)
    if not rows:
        return {}, set(), []

    header_idx = _find_header_row(rows, ALIASES["assembly"] + ALIASES["part_code"])
    col_map = _build_col_map(rows[header_idx])
    if "assembly" not in col_map or "part_code" not in col_map:
        return {}, set(), ["Il file assemblaggi non contiene le colonne Assemb. e Parte"]

    parents_by_part: dict[str, deque[str]] = defaultdict(deque)
    assemblies: set[str] = set()
    warnings: list[str] = []
    current_assembly: str | None = None
    current_assembly_qty = 0

    for row in rows[header_idx + 1:]:
        assembly = _str_cell(row, col_map, "assembly")
        part = _str_cell(row, col_map, "part_code")
        qty = _float_cell(row, col_map, "qty")

        if _is_valid_part_code(assembly) and not part:
            current_assembly = assembly
            current_assembly_qty = max(1, int(round(qty or 1)))
            assemblies.add(assembly)
            continue

        if not _is_valid_part_code(part):
            continue
        if not current_assembly:
            warnings.append(f"Parte senza assemblato: {part}")
            continue

        qty_each = max(1, int(round(qty or 1)))
        parents_by_part[part].extend(
            [current_assembly] * (current_assembly_qty * qty_each)
        )

    return dict(parents_by_part), assemblies, warnings


def parse_commessa_files(
    lista_pezzi_path: Path,
    assemblaggi_path: Path | None = None,
) -> tuple[list[dict], dict]:
    """Legge i file iniziali di una commessa.

    ``lista_pezzi_path`` è obbligatorio e determina posizioni, quantità e dati
    tecnici. ``assemblaggi_path`` è opzionale e aggiunge la relazione tra ogni
    pezzo fisico e il relativo assemblato senza modificare le quantità.
    """
    detected_lista = _detect_file_type(lista_pezzi_path)
    if detected_lista != "lavorazioni":
        raise ValueError(
            "Il file Lista pezzi non è una 'Lavorazione per posizione' riconoscibile"
        )

    items, _ = _parse_single(lista_pezzi_path)
    if not items:
        raise ValueError("Il file Lista pezzi non contiene posizioni riconoscibili")

    parent_queues: dict[str, deque[str]] = {}
    assembly_headers: set[str] = set()
    hierarchy_warnings: list[str] = []
    detected_assemblaggi = None

    if assemblaggi_path is not None:
        detected_assemblaggi = _detect_file_type(assemblaggi_path)
        if detected_assemblaggi != "assemblaggi":
            raise ValueError(
                "Il file Pezzi e assemblati non contiene una gerarchia riconoscibile"
            )
        parent_queues, assembly_headers, hierarchy_warnings = _parse_assembly_hierarchy(
            assemblaggi_path
        )

    for item in items:
        pc = item.get("part_code") or ""
        queue = parent_queues.get(pc)
        item["assembly_parent"] = queue.popleft() if queue else None

    unmatched_hierarchy = sum(len(queue) for queue in parent_queues.values())
    missing_hierarchy = sum(
        1 for item in items
        if assemblaggi_path is not None and item.get("assembly_parent") is None
    )

    for item in items:
        item["tipo_profilo"] = None

    report = _validate(items, assembly_headers or None)
    report["detected"] = {
        "lista_pezzi": detected_lista,
        "assemblaggi": detected_assemblaggi,
        "lista_pezzi_filename": lista_pezzi_path.name,
        "assemblaggi_filename": assemblaggi_path.name if assemblaggi_path else None,
    }
    report["positions"] = len({i["part_code"] for i in items if i.get("part_code")})
    report["assemblies"] = len(assembly_headers)
    report["hierarchy"] = {
        "linked_pieces": len(items) - missing_hierarchy if assemblaggi_path else 0,
        "pieces_without_parent": missing_hierarchy if assemblaggi_path else 0,
        "extra_hierarchy_links": unmatched_hierarchy,
    }

    errors = list(report.get("errors", []))
    warnings = list(report.get("warnings", []))
    notes = list(report.get("notes", [])) + hierarchy_warnings
    if unmatched_hierarchy:
        notes.append(
            f"{unmatched_hierarchy} riferimenti del file assemblati non sono presenti nella Lista pezzi"
        )
    if missing_hierarchy:
        notes.append(
            f"{missing_hierarchy} pezzi senza assemblato: saranno gestiti come pezzi sciolti"
        )
    report["errors"] = errors
    report["warnings"] = warnings
    report["notes"] = notes
    report["ok"] = not errors
    report["summary"] = (
        f"Import {'OK' if report['ok'] else 'CON ERRORI'} — "
        f"{len(items)} pezzi fisici · {report['positions']} posizioni · "
        f"{len(assembly_headers)} assemblaggi · "
        f"{len(errors)} errori · {len(warnings)} warning · {len(notes)} note"
    )
    return items, report


# ── Parser unificato due file (compatibilità) ─────────────────────────────────

def parse_two_files(
    file_a: Path,
    file_b: Path,
) -> tuple[list[dict], dict]:
    """Compatibilità per i chiamanti storici; usa il parser commessa corretto."""
    type_a = _detect_file_type(file_a)
    type_b = _detect_file_type(file_b)

    if type_a == "lavorazioni" and type_b == "assemblaggi":
        return parse_commessa_files(file_a, file_b)
    if type_b == "lavorazioni" and type_a == "assemblaggi":
        return parse_commessa_files(file_b, file_a)
    raise ValueError("Servono una Lista pezzi e una Lista pezzi e assemblati riconoscibili")


# ── Validazione ───────────────────────────────────────────────────────────────

def _validate(items: list[dict], known_assemblies: set[str] | None = None) -> dict:
    errors: list[str]   = []
    warnings: list[str] = []

    # Tutti i codici validi come parent: part_codes + assemblies dichiarati
    part_codes = {i["part_code"] for i in items if i["part_code"]}
    valid_parents = (known_assemblies or part_codes) | part_codes

    seen_instance_keys: set[str] = set()

    for item in items:
        pc = item["part_code"]
        inst = item.get("instance_number", 1)
        key = f"{pc}-{inst:03d}"

        if key in seen_instance_keys:
            errors.append(f"Posizione duplicata: {key}")
        else:
            seen_instance_keys.add(key)

        if not item.get("profile"):
            errors.append(f"Profilo mancante: {pc}")

        if item.get("assembly_parent") and item["assembly_parent"] not in valid_parents:
            errors.append(f"Assemblaggio orfano: {pc} → {item['assembly_parent']}")

        if not item.get("qty") or item["qty"] <= 0:
            errors.append(f"Qty zero: {pc}")

        if item.get("weight_kg") is None:
            warnings.append(f"Peso nullo: {pc}")

    unique_parts     = len({i["part_code"] for i in items})
    unique_asm       = len({i["assembly_parent"] for i in items if i.get("assembly_parent")})
    physical_pieces  = len(items)

    ok = len(errors) == 0
    summary = (
        f"Import {'OK' if ok else 'CON ERRORI'} — "
        f"{physical_pieces} pezzi fisici · {unique_parts} posizioni · "
        f"{unique_asm} assemblaggi · {len(errors)} errori · {len(warnings)} warning"
    )

    return {
        "ok":             ok,
        "summary":        summary,
        "total_pieces":   physical_pieces,
        "unique_parts":   unique_parts,
        "assemblies":     unique_asm,
        "errors":         errors,
        "warnings":       warnings,
    }


# ── Backward compat ───────────────────────────────────────────────────────────

def parse_distinta_file(file_path: Path) -> list[dict]:
    """Legacy: singolo file generico. Converte l'output normalizzato nel formato
    atteso da create_distinta_item (part_number, description, material_code, …)."""
    items, _ = _parse_single(file_path)
    return [_normalized_to_db(i) for i in items]


def parse_lista_parti(file_path: Path) -> list[dict]:
    """Legacy: parser specifico per 'Lista parti assemblaggi'."""
    return parse_distinta_file(file_path)


def normalized_to_db_bulk(items: list[dict]) -> list[dict]:
    """Converte una lista di item normalizzati nel formato DB (part_number, description, …)."""
    return [_normalized_to_db(i) for i in items]


def _normalized_to_db(item: dict) -> dict:
    return {
        "part_number":          item.get("part_code"),
        "description":          item.get("profile"),
        "quantity":             float(item.get("qty") or 1),
        "material_code":        item.get("material"),
        "material_description": None,
        "commessa_reference":   item.get("commessa_ref"),
        "length_mm":            item.get("length_mm"),
        "width_mm":             item.get("width_mm"),
        "weight_kg":            item.get("weight_kg"),
        "instance_number":      item.get("instance_number"),
        "parent_assembly":      item.get("assembly_parent"),
        "tipo_profilo":         item.get("tipo_profilo"),
    }
