"""Parser per il file Fasi Operative (.xlsx/.xlsm) e calcolo percorso pezzi.

Struttura attesa:
  - Riga 4: intestazione colonne
  - Righe 5+: dati

Colonne (0-based):
  0  Marca/Pos.
  1  Profilo
  2  Q.tà
  3  Fase
  4  Postazione
  5  Tempo prev. (min/pz)
  6  Tempo tot. (min)
  7  Sequenza
  8  Dipende da
  9  Note operative
"""
import json
from collections import defaultdict, deque
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List


def parse_fasi_operative(file_path: Path) -> List[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=file_path, data_only=True, keep_vba=False)
    sheet = wb.active

    def _str(v) -> str | None:
        return str(v).strip() if v not in (None, "") else None

    def _float(v) -> float | None:
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _int(v) -> int | None:
        try:
            return int(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    rows = list(sheet.iter_rows(min_row=5, values_only=True))  # skip header at row 4
    items: List[dict] = []

    for row in rows:
        if not row or len(row) < 4:
            continue
        # Skip completely empty rows
        if all(v in (None, "") for v in row):
            continue

        fase = _str(row[3]) if len(row) > 3 else None
        if not fase:
            continue  # skip rows without a fase name

        items.append({
            "marca_pos":       _str(row[0]) if len(row) > 0 else None,
            "profilo":         _str(row[1]) if len(row) > 1 else None,
            "quantita":        _float(row[2]) if len(row) > 2 else None,
            "fase":            fase,
            "postazione":      _str(row[4]) if len(row) > 4 else None,
            "tempo_prev_minpz": _float(row[5]) if len(row) > 5 else None,
            "tempo_tot_min":   _float(row[6]) if len(row) > 6 else None,
            "sequenza":        _int(row[7]) if len(row) > 7 else None,
            "dipende_da":      _str(row[8]) if len(row) > 8 else None,
            "note_operative":  _str(row[9]) if len(row) > 9 else None,
        })

    return items


def compute_percorsi(fasi_rows: list, commessa_id: int) -> List[Dict[str, Any]]:
    """
    Data la lista di FaseOperativa ORM objects (dopo il flush),
    raggruppa per marca_pos e calcola il percorso ordinato per ciascun pezzo.

    Ordinamento: Kahn topological sort su dipende_da (contiene sequenza IDs
    separati da virgola). Fallback a sort per sequenza se il grafo è aciclico
    o se dipende_da è assente.

    Restituisce una lista di dict pronti per l'upsert su pezzo_percorso:
      [{"commessa_id": ..., "marca_pos": ..., "percorso": "<JSON>"}]
    """
    # Raggruppa per marca_pos (skippa righe senza marca)
    by_marca: Dict[str, list] = defaultdict(list)
    for f in fasi_rows:
        if f.marca_pos:
            by_marca[f.marca_pos].append(f)

    result = []
    for marca_pos, fasi in by_marca.items():
        # Mappa sequenza → fase (per risolvere dipende_da)
        seq_map: Dict[int, Any] = {f.sequenza: f for f in fasi if f.sequenza is not None}

        # Costruisci grafo di dipendenze: seq → set(seq_predecessori)
        deps: Dict[int, set] = {}
        for f in fasi:
            if f.sequenza is None:
                continue
            predecessori = set()
            if f.dipende_da:
                for tok in str(f.dipende_da).replace(";", ",").split(","):
                    tok = tok.strip()
                    try:
                        predecessori.add(int(float(tok)))
                    except ValueError:
                        pass
            deps[f.sequenza] = predecessori

        # Kahn's algorithm per topological sort
        in_degree = {s: len(d) for s, d in deps.items()}
        queue = deque(s for s, d in in_degree.items() if d == 0)
        sorted_seqs: List[int] = []

        while queue:
            # Tra i nodi disponibili, prendi quello con sequenza minore
            queue = deque(sorted(queue))
            node = queue.popleft()
            sorted_seqs.append(node)
            for s, predecessori in deps.items():
                if node in predecessori:
                    in_degree[s] -= 1
                    if in_degree[s] == 0:
                        queue.append(s)

        # Se ci sono sequenze non risolte (ciclo o senza sequenza), appendile ordinate
        resolved = set(sorted_seqs)
        extra = sorted(s for s in deps if s not in resolved)
        sorted_seqs.extend(extra)

        # Aggiungi fasi senza sequenza in fondo
        no_seq = [f for f in fasi if f.sequenza is None]

        # Costruisci array percorso
        path = []
        for seq in sorted_seqs:
            if seq not in seq_map:
                continue
            f = seq_map[seq]
            path.append({
                "fase_id":          f.id,
                "sequenza":         f.sequenza,
                "fase":             f.fase,
                "postazione":       f.postazione,
                "dipende_da":       f.dipende_da,
                "tempo_prev_minpz": float(f.tempo_prev_minpz) if f.tempo_prev_minpz is not None else None,
                "tempo_tot_min":    float(f.tempo_tot_min)    if f.tempo_tot_min    is not None else None,
                "note_operative":   f.note_operative,
            })
        for f in no_seq:
            path.append({
                "fase_id":          f.id,
                "sequenza":         None,
                "fase":             f.fase,
                "postazione":       f.postazione,
                "dipende_da":       f.dipende_da,
                "tempo_prev_minpz": float(f.tempo_prev_minpz) if f.tempo_prev_minpz is not None else None,
                "tempo_tot_min":    float(f.tempo_tot_min)    if f.tempo_tot_min    is not None else None,
                "note_operative":   f.note_operative,
            })

        result.append({
            "commessa_id": commessa_id,
            "marca_pos":   marca_pos,
            "percorso":    json.dumps(path, ensure_ascii=False),
        })

    return result
