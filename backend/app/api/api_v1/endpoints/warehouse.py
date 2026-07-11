from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import List, Literal

import base64
import csv
import io
import json
import os
import re
from pathlib import Path
from tempfile import NamedTemporaryFile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from fpdf import FPDF
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from backend.app.crud import warehouse as crud
from backend.app.db.session import get_db
from backend.app.models.warehouse import (
    Batch,
    Certificate,
    Material,
    MovementType,
    Receipt,
    StockMovement,
    WarehouseCustomField,
    WarehouseCustomValue,
    WarehouseItem,
)
from backend.app.schemas import warehouse as warehouse_schemas
from backend.app.services.qr import generate_qr_for_payload, generate_qr_for_uuid
from backend.app.services.warehouse_items import (
    close_items_for_outgoing,
    create_items_for_incoming,
)

router = APIRouter()


@router.post("/suppliers", response_model=warehouse_schemas.SupplierRead)
def create_supplier(
    supplier_in: warehouse_schemas.SupplierCreate,
    db: Session = Depends(get_db),
):
    return crud.create_supplier(db=db, obj_in=supplier_in)


@router.get("/suppliers", response_model=List[warehouse_schemas.SupplierRead])
def list_suppliers(skip: int = 0, limit: int = 100, q: str | None = None, db: Session = Depends(get_db)):
    return crud.get_suppliers(db=db, skip=skip, limit=limit, q=q)


@router.get("/suppliers/{supplier_id}", response_model=warehouse_schemas.SupplierRead)
def get_supplier(supplier_id: int, db: Session = Depends(get_db)):
    supplier = crud.get_supplier(db=db, supplier_id=supplier_id)
    if supplier is None:
        raise HTTPException(status_code=404, detail="Fornitore non trovato")
    return supplier


@router.post("/materials", response_model=warehouse_schemas.MaterialRead)
def create_material(
    material_in: warehouse_schemas.MaterialCreate,
    db: Session = Depends(get_db),
):
    return crud.create_material(db=db, obj_in=material_in)


def _material_code_from_input(material_in: warehouse_schemas.MaterialIncomingCreate) -> str:
    if material_in.code and material_in.code.strip():
        return material_in.code.strip().upper()
    parts = [
        material_in.tipo,
        material_in.profilo,
        material_in.dimensioni,
        material_in.qualita,
        material_in.colata,
    ]
    code = "-".join(str(part).strip().upper().replace(" ", "") for part in parts if part)
    if not code:
        raise HTTPException(status_code=422, detail="Codice o almeno tipo/profilo obbligatori")
    return code


@router.post("/materials/with-incoming", response_model=warehouse_schemas.MagazzinoItemRead)
def create_material_with_incoming(
    material_in: warehouse_schemas.MaterialIncomingCreate,
    db: Session = Depends(get_db),
):
    row = _upsert_material_with_incoming(db, material_in, allow_existing=False)
    db.commit()
    row.pop("_created", None)
    row.pop("_physical_items_created", None)
    return row


def _upsert_material_with_incoming(
    db: Session,
    material_in: warehouse_schemas.MaterialIncomingCreate,
    allow_existing: bool = True,
    read_back: bool = True,
) -> dict:
    code = _material_code_from_input(material_in)
    existing = db.scalar(select(Material).where(Material.code == code))
    if existing is not None and not allow_existing:
        raise HTTPException(status_code=409, detail="Codice materiale già presente: usa Ingresso su materiale esistente")

    description = material_in.description or " · ".join(
        part for part in [material_in.tipo, material_in.profilo, material_in.dimensioni, material_in.qualita] if part
    ) or code
    created = existing is None
    if existing is None:
        material = Material(
            code=code,
            description=description,
            unit=material_in.unit or "PZ",
            specification=material_in.specification,
            tipo=material_in.tipo,
            profilo=material_in.profilo,
            dimensioni=material_in.dimensioni,
            norma_uni=material_in.norma_uni,
            qualita=material_in.qualita,
            colata=material_in.colata,
            commessa_ref=material_in.commessa_ref,
            uso_materiale=material_in.uso_materiale,
            posizione=material_in.posizione,
            peso_u_kg=material_in.peso_u_kg,
            peso_1_pz=material_in.peso_1_pz,
        )
        db.add(material)
        db.flush()
    else:
        material = existing
        material.description = material_in.description or material.description
        material.unit = material_in.unit or material.unit
        material.specification = material_in.specification or material.specification
        material.tipo = material_in.tipo or material.tipo
        material.profilo = material_in.profilo or material.profilo
        material.dimensioni = material_in.dimensioni or material.dimensioni
        material.norma_uni = material_in.norma_uni or material.norma_uni
        material.qualita = material_in.qualita or material.qualita
        material.colata = material_in.colata or material.colata
        material.commessa_ref = material_in.commessa_ref or material.commessa_ref
        material.uso_materiale = material_in.uso_materiale or material.uso_materiale
        material.posizione = material_in.posizione or material.posizione
        material.peso_u_kg = material_in.peso_u_kg or material.peso_u_kg
        material.peso_1_pz = material_in.peso_1_pz or material.peso_1_pz
    physical_created = 0
    movement_created = 0
    reserved_created = 0
    if float(material_in.quantity or 0) > 0:
        movement = StockMovement(
            material_id=material.id,
            quantity=material_in.quantity,
            movement_type=MovementType.INCOMING,
            reason=material_in.reason or "Ingresso nuovo materiale",
        )
        db.add(movement)
        db.flush()
        movement_created = 1
        try:
            new_items = create_items_for_incoming(db, material.id, material_in.quantity, movement.id)
            physical_created = len(new_items)
            reservation = _reservation_or_none(material_in.commessa_ref)
            if reservation:
                now = datetime.utcnow()
                for warehouse_item in new_items:
                    warehouse_item.status = "RESERVED"
                    warehouse_item.reserved_for_commessa = reservation
                    warehouse_item.reserved_at = now
                reserved_created = len(new_items)
        except ValueError as exc:
            db.rollback()
            raise HTTPException(status_code=422, detail=str(exc))
    if read_back:
        row = next((r for r in crud.get_magazzino_list(db=db, limit=250, q=code) if r["material_id"] == material.id), None)
        if row is None:
            row = {
                "material_id": material.id,
                "material_code": material.code,
                "tipo": material.tipo,
                "profilo": material.profilo,
                "n_pezzi": float(material_in.quantity),
                "dimensioni": material.dimensioni,
                "qualita": material.qualita,
                "colata": material.colata,
                "commessa_ref": material.commessa_ref,
                "uso_materiale": material.uso_materiale,
                "posizione": material.posizione,
                "peso_kg": None,
                "peso_u_kg": float(material.peso_u_kg) if material.peso_u_kg is not None else None,
                "peso_1_pz": float(material.peso_1_pz) if material.peso_1_pz is not None else None,
                "norma_uni": material.norma_uni,
                "physical_items_count": physical_created,
                "reserved_items_count": 0,
                "reserved_commesse": [],
                "custom_fields": {},
            }
    else:
        row = {"material_id": material.id}
    row["_created"] = created
    row["_physical_items_created"] = physical_created
    row["_movement_created"] = movement_created
    row["_reserved_items_created"] = reserved_created
    row["_material_id"] = material.id
    return row


@router.post("/materials/bulk-incoming", response_model=warehouse_schemas.MaterialIncomingBulkResult)
def create_materials_bulk_incoming(
    request: warehouse_schemas.MaterialIncomingBulkCreate,
    db: Session = Depends(get_db),
):
    created = 0
    existing = 0
    physical = 0
    movements = 0
    reserved = 0
    for item in request.items:
        row = _upsert_material_with_incoming(db, item, allow_existing=True, read_back=False)
        if row.get("_created"):
            created += 1
        else:
            existing += 1
        physical += int(row.get("_physical_items_created") or item.quantity or 0)
        movements += int(row.get("_movement_created") or 0)
        reserved += int(row.get("_reserved_items_created") or 0)
    db.commit()
    return {
        "rows": len(request.items),
        "materials_created": created,
        "materials_existing": existing,
        "movements_created": movements,
        "physical_items_created": physical,
        "reserved_items_created": reserved,
    }


def _normalize_import_key(value: object) -> str:
    key = str(value or "").strip().lower()
    key = key.replace("à", "a").replace("è", "e").replace("é", "e").replace("ì", "i").replace("ò", "o").replace("ù", "u")
    key = re.sub(r"[^a-z0-9]+", "_", key).strip("_")
    aliases = {
        "codice": "code",
        "material_code": "code",
        "descrizione": "description",
        "description": "description",
        "qta": "quantity",
        "qty": "quantity",
        "quantita": "quantity",
        "n_pezzi": "quantity",
        "n_pezzi_attuale": "quantity",
        "pezzi_attuale": "quantity",
        "stock_attuale": "quantity",
        "giacenza": "quantity",
        "pezzi": "quantity",
        "unita": "unit",
        "um": "unit",
        "norma": "norma_uni",
        "qualita": "qualita",
        "qualita_": "qualita",
        "peso_u": "peso_u_kg",
        "peso_u_kg": "peso_u_kg",
        "peso_pezzo": "peso_1_pz",
        "peso_1_pz": "peso_1_pz",
        "uso": "uso_materiale",
        "uso_materiale": "uso_materiale",
        "tipologia": "uso_materiale",
        "posizione": "posizione",
        "ubicazione": "posizione",
        "location": "posizione",
        "commessa": "commessa_ref",
        "commessa_ref": "commessa_ref",
        "prenotazione": "commessa_ref",
        "prenotato_per": "commessa_ref",
        "causale": "reason",
    }
    return aliases.get(key, key)


KNOWN_IMPORT_KEYS = {
    "code",
    "description",
    "quantity",
    "unit",
    "tipo",
    "profilo",
    "dimensioni",
    "norma_uni",
    "qualita",
    "colata",
    "uso_materiale",
    "posizione",
    "commessa_ref",
    "peso_u_kg",
    "peso_1_pz",
    "reason",
}


def _value(row: dict, *keys: str):
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def _float_or_none(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _physical_quantity(value: float) -> int:
    """L'inventario fisico genera un QR per elemento: la quantità deve essere intera."""
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _text_or_none(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _reservation_or_none(value: object) -> str | None:
    text = _text_or_none(value)
    if not text:
        return None
    normalized = re.sub(r"[^A-Z0-9]+", "", text.upper())
    if normalized in {"MAG", "MG", "MAGAZZINO", "STOCK", "SCORTE"}:
        return None
    cleaned = text.upper().strip()
    cleaned = re.sub(r"^(COMMESSA|COMM)\s*", "", cleaned).strip()
    cleaned = re.sub(r"[^A-Z0-9]+", "_", cleaned).strip("_")
    return cleaned or text


def _clean_import_value(value: object) -> str | None:
    if value in (None, ""):
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _custom_label_from_key(key: str) -> str:
    return " ".join(part.capitalize() for part in key.split("_") if part) or key


def _upsert_custom_values(
    db: Session,
    material_id: int,
    custom_values: dict[str, str],
    custom_labels: dict[str, str],
) -> None:
    for key, raw_value in custom_values.items():
        value = _clean_import_value(raw_value)
        if value is None:
            continue
        field = db.scalar(select(WarehouseCustomField).where(WarehouseCustomField.key == key))
        if field is None:
            field = WarehouseCustomField(
                key=key,
                label=(custom_labels.get(key) or _custom_label_from_key(key))[:200],
                value_type="text",
                created_at=datetime.utcnow(),
            )
            db.add(field)
            db.flush()
        elif custom_labels.get(key) and field.label == _custom_label_from_key(field.key):
            field.label = custom_labels[key][:200]
        current = db.scalar(
            select(WarehouseCustomValue).where(
                WarehouseCustomValue.material_id == material_id,
                WarehouseCustomValue.field_id == field.id,
            )
        )
        if current is None:
            current = WarehouseCustomValue(material_id=material_id, field_id=field.id)
            db.add(current)
        current.value = value


def _custom_fields_for_material_ids(db: Session, material_ids: list[int]) -> dict[int, dict[str, str]]:
    if not material_ids:
        return {}
    result: dict[int, dict[str, str]] = {int(material_id): {} for material_id in material_ids}
    rows = db.execute(
        select(WarehouseCustomValue.material_id, WarehouseCustomField.key, WarehouseCustomValue.value)
        .join(WarehouseCustomField, WarehouseCustomField.id == WarehouseCustomValue.field_id)
        .where(WarehouseCustomValue.material_id.in_(material_ids))
        .order_by(WarehouseCustomField.label)
    ).all()
    for material_id, key, value in rows:
        if value not in (None, ""):
            result.setdefault(int(material_id), {})[str(key)] = str(value)
    return result


def _rows_to_incoming_with_custom(
    rows: list[dict],
    reason: str,
) -> list[tuple[warehouse_schemas.MaterialIncomingCreate, dict[str, str], dict[str, str]]]:
    parsed: list[tuple[warehouse_schemas.MaterialIncomingCreate, dict[str, str], dict[str, str]]] = []
    for raw in rows:
        row: dict[str, object] = {}
        custom_values: dict[str, str] = {}
        custom_labels: dict[str, str] = {}
        for raw_key, value in raw.items():
            normalized = _normalize_import_key(raw_key)
            if not normalized:
                continue
            if normalized in KNOWN_IMPORT_KEYS:
                row[normalized] = value
            else:
                cleaned = _clean_import_value(value)
                if cleaned is not None:
                    row[normalized] = value
                    custom_values[normalized] = cleaned
                    custom_labels[normalized] = str(raw_key or normalized).strip() or _custom_label_from_key(normalized)
        raw_quantity = _float_or_none(_value(row, "quantity"))
        if raw_quantity is None or raw_quantity < 0:
            continue
        if not _value(row, "code") and not (_value(row, "tipo") and _value(row, "profilo")):
            continue
        quantity = _physical_quantity(raw_quantity)
        if abs(float(raw_quantity) - quantity) > 1e-9:
            custom_values.setdefault("quantita_file_originale", f"{raw_quantity:g}")
            custom_labels.setdefault("quantita_file_originale", "Quantità file originale")
        payload = {
            "code": _text_or_none(_value(row, "code")),
            "description": _text_or_none(_value(row, "description")),
            "unit": _text_or_none(_value(row, "unit")) or "PZ",
            "tipo": _text_or_none(_value(row, "tipo")),
            "profilo": _text_or_none(_value(row, "profilo")),
            "dimensioni": _text_or_none(_value(row, "dimensioni")),
            "norma_uni": _text_or_none(_value(row, "norma_uni")),
            "qualita": _text_or_none(_value(row, "qualita")),
            "colata": _text_or_none(_value(row, "colata")),
            "uso_materiale": _text_or_none(_value(row, "uso_materiale", "uso", "tipologia")),
            "posizione": _text_or_none(_value(row, "posizione", "ubicazione", "location")),
            "commessa_ref": _reservation_or_none(_value(row, "commessa_ref")) or _text_or_none(_value(row, "commessa_ref")),
            "peso_u_kg": _float_or_none(_value(row, "peso_u_kg")),
            "peso_1_pz": _float_or_none(_value(row, "peso_1_pz")),
            "quantity": quantity,
            "reason": _text_or_none(_value(row, "reason")) or reason,
        }
        parsed.append((warehouse_schemas.MaterialIncomingCreate.model_validate(payload), custom_values, custom_labels))
    return parsed


def _rows_to_incoming(rows: list[dict], reason: str) -> list[warehouse_schemas.MaterialIncomingCreate]:
    return [item for item, _, _ in _rows_to_incoming_with_custom(rows, reason)]


def _parse_append_file(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(4096)
            fh.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\\t")
            return list(csv.DictReader(fh, dialect=dialect))
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows, [])]
        return [dict(zip(headers, row)) for row in rows if any(v not in (None, "") for v in row)]
    if suffix == ".pdf":
        import pdfplumber

        parsed: list[dict] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    if not table or len(table) < 2:
                        continue
                    headers = [str(value or "").strip() for value in table[0]]
                    for row in table[1:]:
                        if any(v not in (None, "") for v in row):
                            parsed.append(dict(zip(headers, row)))
        return parsed
    raise HTTPException(status_code=422, detail="Formato non supportato. Usa CSV, XLSX/XLSM o PDF tabellare.")


@router.post("/materials/import-append", response_model=warehouse_schemas.MaterialIncomingBulkResult)
async def import_append_materials(
    file: UploadFile = File(...),
    reason: str = Query("Import append magazzino"),
    db: Session = Depends(get_db),
):
    suffix = Path(file.filename or "").suffix.lower() or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = Path(tmp.name)
    try:
        rows = _parse_append_file(tmp_path)
        parsed_items = _rows_to_incoming_with_custom(rows, reason)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Impossibile leggere il file: {exc}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
    if not parsed_items:
        raise HTTPException(status_code=422, detail="Nessuna riga valida trovata. Servono almeno codice oppure tipo/profilo e una quantità attuale valida anche zero.")
    created = 0
    existing = 0
    physical = 0
    movements = 0
    reserved = 0
    for item, custom_values, custom_labels in parsed_items:
        row = _upsert_material_with_incoming(db, item, allow_existing=True, read_back=False)
        if row.get("_created"):
            created += 1
        else:
            existing += 1
        physical += int(row.get("_physical_items_created") or item.quantity or 0)
        movements += int(row.get("_movement_created") or 0)
        reserved += int(row.get("_reserved_items_created") or 0)
        material_id = int(row.get("_material_id") or row["material_id"])
        _upsert_custom_values(db, material_id, custom_values, custom_labels)
    db.commit()
    return {
        "rows": len(parsed_items),
        "materials_created": created,
        "materials_existing": existing,
        "movements_created": movements,
        "physical_items_created": physical,
        "reserved_items_created": reserved,
    }


@router.get("/materials", response_model=List[warehouse_schemas.MaterialRead])
def list_materials(
    skip: int = 0,
    limit: int = 100,
    q: str | None = None,
    code: str | None = None,
    db: Session = Depends(get_db),
):
    return crud.get_materials(db=db, skip=skip, limit=limit, q=q, code=code)


@router.get("/materials/{material_id}", response_model=warehouse_schemas.MaterialRead)
def get_material(material_id: int, db: Session = Depends(get_db)):
    material = crud.get_material(db=db, material_id=material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Materiale non trovato")
    return material


@router.post("/batches", response_model=warehouse_schemas.BatchRead)
def create_batch(
    batch_in: warehouse_schemas.BatchCreate,
    db: Session = Depends(get_db),
):
    return crud.create_batch(db=db, obj_in=batch_in)


@router.get("/batches", response_model=List[warehouse_schemas.BatchRead])
def list_batches(skip: int = 0, limit: int = 100, material_id: int | None = None, db: Session = Depends(get_db)):
    return crud.get_batches(db=db, skip=skip, limit=limit, material_id=material_id)


@router.get("/batches/{batch_id}", response_model=warehouse_schemas.BatchRead)
def get_batch(batch_id: int, db: Session = Depends(get_db)):
    batch = crud.get_batch(db=db, batch_id=batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Lotto non trovato")
    return batch


@router.post("/receipts", response_model=warehouse_schemas.ReceiptRead)
def create_receipt(
    receipt_in: warehouse_schemas.ReceiptCreate,
    db: Session = Depends(get_db),
):
    if receipt_in.quantity <= 0:
        raise HTTPException(status_code=422, detail="La quantità deve essere maggiore di zero")
    return crud.create_receipt(db=db, obj_in=receipt_in)


@router.get("/receipts", response_model=List[warehouse_schemas.ReceiptRead])
def list_receipts(
    skip: int = 0,
    limit: int = 50,
    supplier_id: int | None = None,
    material_id: int | None = None,
    db: Session = Depends(get_db),
):
    return crud.get_receipts(db=db, skip=skip, limit=limit, supplier_id=supplier_id, material_id=material_id)


@router.get("/movements", response_model=List[warehouse_schemas.StockMovementRead])
def list_movements(
    skip: int = 0,
    limit: int = 50,
    material_id: int | None = None,
    commessa_id: int | None = None,
    movement_type: str | None = None,
    db: Session = Depends(get_db),
):
    if commessa_id is not None:
        raise HTTPException(
            status_code=410,
            detail="Il magazzino e le commesse sono separati: i movimenti non sono filtrabili per commessa.",
        )
    return crud.get_movements(db=db, skip=skip, limit=limit, material_id=material_id, movement_type=movement_type)


@router.post("/movements", response_model=warehouse_schemas.StockMovementRead)
def create_stock_movement(
    movement_in: warehouse_schemas.StockMovementCreate,
    db: Session = Depends(get_db),
):
    if getattr(movement_in, "commessa_id", None) is not None or getattr(movement_in, "destination_commessa", None):
        raise HTTPException(
            status_code=422,
            detail="Il movimento di magazzino non può essere collegato a una commessa in questa fase.",
        )
    if (
        movement_in.movement_type == warehouse_schemas.MovementType.ADJUSTMENT
        and movement_in.quantity == 0
    ):
        raise HTTPException(status_code=422, detail="La rettifica non può essere zero")
    if (
        movement_in.movement_type != warehouse_schemas.MovementType.ADJUSTMENT
        and movement_in.quantity <= 0
    ):
        raise HTTPException(status_code=422, detail="La quantità deve essere maggiore di zero")
    movement = StockMovement(**movement_in.model_dump())
    db.add(movement)
    db.flush()
    try:
        if movement_in.movement_type == warehouse_schemas.MovementType.INCOMING:
            create_items_for_incoming(db, movement.material_id, movement.quantity, movement.id)
        elif movement_in.movement_type in (
            warehouse_schemas.MovementType.OUTGOING,
            warehouse_schemas.MovementType.SFRIDO,
        ):
            close_items_for_outgoing(db, movement.material_id, movement.quantity, movement.id)
        elif movement_in.movement_type == warehouse_schemas.MovementType.ADJUSTMENT:
            if movement.quantity > 0:
                create_items_for_incoming(db, movement.material_id, movement.quantity, movement.id)
            elif movement.quantity < 0:
                close_items_for_outgoing(db, movement.material_id, abs(movement.quantity), movement.id)
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=422, detail=str(exc))
    db.commit()
    db.refresh(movement)
    return movement


@router.get("/stock/{material_id}", response_model=warehouse_schemas.StockBalanceRead)
def read_stock_balance(material_id: int, db: Session = Depends(get_db)):
    balance = crud.get_stock_balance(db=db, material_id=material_id)
    if balance is None:
        raise HTTPException(status_code=404, detail="Materiale non trovato")
    return balance


@router.get("/magazzino", response_model=List[warehouse_schemas.MagazzinoItemRead])
def list_magazzino(
    skip: int = 0,
    limit: int = 200,
    q: str | None = None,
    db: Session = Depends(get_db),
):
    return crud.get_magazzino_list(db=db, skip=skip, limit=limit, q=q)


@router.get("/custom-fields", response_model=List[warehouse_schemas.WarehouseCustomFieldRead])
def list_warehouse_custom_fields(db: Session = Depends(get_db)):
    return db.scalars(select(WarehouseCustomField).order_by(WarehouseCustomField.label)).all()


@router.get("/export.csv")
def export_warehouse_csv(db: Session = Depends(get_db)):
    custom_fields = db.scalars(select(WarehouseCustomField).order_by(WarehouseCustomField.label)).all()
    rows = crud.get_magazzino_list(db=db, limit=100000)
    output = io.StringIO()
    base_headers = [
        "material_id",
        "material_code",
        "tipo",
        "uso_materiale",
        "posizione",
        "profilo",
        "dimensioni",
        "norma_uni",
        "qualita",
        "colata",
        "n_pezzi",
        "commessa",
        "peso_kg",
        "peso_u_kg",
        "peso_1_pz",
        "physical_items_count",
        "reserved_items_count",
    ]
    custom_headers = [f"custom:{field.key}" for field in custom_fields]
    writer = csv.DictWriter(output, fieldnames=base_headers + custom_headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        csv_row = {key: row.get(key) for key in base_headers}
        csv_row["commessa"] = ", ".join(row.get("reserved_commesse") or [])
        for field in custom_fields:
            csv_row[f"custom:{field.key}"] = (row.get("custom_fields") or {}).get(field.key, "")
        writer.writerow(csv_row)
    return Response(
        content=output.getvalue().encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="inventario_magazzino.csv"'},
    )


def _physical_item_read(item: WarehouseItem) -> dict:
    material = item.material
    effective_status = "RESERVED" if item.reserved_for_commessa and item.status == "AVAILABLE" else item.status
    return {
        "id": item.id,
        "uuid": item.uuid,
        "material_id": item.material_id,
        "ordinal": item.ordinal,
        "status": effective_status,
        "created_at": item.created_at,
        "exited_at": item.exited_at,
        "label": f"{material.code} · #{item.ordinal:04d}",
        "reserved_for_commessa": item.reserved_for_commessa,
        "uso_materiale": item.uso_materiale or material.uso_materiale,
        "posizione": item.posizione or material.posizione,
    }


def _decimal_or_none(value: object) -> float | None:
    return float(value) if value is not None else None


def _item_value(item: WarehouseItem, material: Material, field: str):
    value = getattr(item, field, None)
    if value is not None:
        return value
    return getattr(material, field, None)


def _item_detail_read(item: WarehouseItem, db: Session) -> dict:
    material = item.material
    peso_1_pz = _item_value(item, material, "peso_1_pz")
    manual_fields = [
        field
        for field in (
            "tipo",
            "profilo",
            "dimensioni",
            "norma_uni",
            "qualita",
            "colata",
            "uso_materiale",
            "posizione",
            "commessa_ref",
            "reserved_for_commessa",
            "peso_u_kg",
            "peso_1_pz",
            "notes",
        )
        if getattr(item, field, None) is not None
    ]
    return {
        **_physical_item_read(item),
        "material_code": material.code,
        "description": material.description,
        "n_pezzi": 1,
        "tipo": _item_value(item, material, "tipo"),
        "profilo": _item_value(item, material, "profilo"),
        "dimensioni": _item_value(item, material, "dimensioni"),
        "norma_uni": _item_value(item, material, "norma_uni"),
        "qualita": _item_value(item, material, "qualita"),
        "colata": _item_value(item, material, "colata"),
        "uso_materiale": _item_value(item, material, "uso_materiale"),
        "posizione": _item_value(item, material, "posizione"),
        "commessa_ref": _item_value(item, material, "commessa_ref"),
        "reserved_for_commessa": item.reserved_for_commessa,
        "peso_u_kg": _decimal_or_none(_item_value(item, material, "peso_u_kg")),
        "peso_1_pz": _decimal_or_none(peso_1_pz),
        "peso_kg": _decimal_or_none(peso_1_pz),
        "unit": material.unit,
        "source_movement_id": item.source_movement_id,
        "exit_movement_id": item.exit_movement_id,
        "notes": item.notes,
        "manual_overrides": manual_fields,
        "custom_fields": _custom_fields_for_material_ids(db, [material.id]).get(material.id, {}),
    }


@router.get(
    "/materials/{material_id}/items",
    response_model=List[warehouse_schemas.WarehousePhysicalItemRead],
)
def list_material_items(
    material_id: int,
    include_closed: bool = False,
    db: Session = Depends(get_db),
):
    if db.get(Material, material_id) is None:
        raise HTTPException(status_code=404, detail="Materiale non trovato")
    stmt = select(WarehouseItem).where(WarehouseItem.material_id == material_id)
    if not include_closed:
        stmt = stmt.where(WarehouseItem.status.in_(["AVAILABLE", "RESERVED"]))
    items = db.scalars(stmt.order_by(WarehouseItem.ordinal)).all()
    return [_physical_item_read(item) for item in items]


@router.get(
    "/items",
    response_model=List[warehouse_schemas.WarehouseItemDetailRead],
)
def list_warehouse_items(
    skip: int = 0,
    limit: int = 5000,
    include_closed: bool = False,
    q: str | None = None,
    db: Session = Depends(get_db),
):
    stmt = (
        select(WarehouseItem)
        .join(WarehouseItem.material)
        .options(joinedload(WarehouseItem.material))
        .order_by(Material.tipo, Material.profilo, Material.code, WarehouseItem.ordinal)
    )
    if not include_closed:
        stmt = stmt.where(WarehouseItem.status.in_(["AVAILABLE", "RESERVED"]))
    if q:
        like = f"%{q}%"
        custom_material_ids = (
            select(WarehouseCustomValue.material_id)
            .where(WarehouseCustomValue.value.ilike(like))
            .distinct()
        )
        stmt = stmt.where(
            Material.code.ilike(like)
            | Material.tipo.ilike(like)
            | Material.profilo.ilike(like)
            | Material.qualita.ilike(like)
            | Material.colata.ilike(like)
            | Material.uso_materiale.ilike(like)
            | Material.posizione.ilike(like)
            | Material.id.in_(custom_material_ids)
            | WarehouseItem.uuid.ilike(like)
            | WarehouseItem.uso_materiale.ilike(like)
            | WarehouseItem.posizione.ilike(like)
            | WarehouseItem.reserved_for_commessa.ilike(like)
        )
    items = db.scalars(stmt.offset(skip).limit(limit)).all()
    return [_item_detail_read(item, db) for item in items]


@router.get(
    "/items/{item_uuid}",
    response_model=warehouse_schemas.WarehouseItemDetailRead,
)
def get_warehouse_item(item_uuid: str, db: Session = Depends(get_db)):
    item = db.scalar(
        select(WarehouseItem)
        .options(joinedload(WarehouseItem.material))
        .where(WarehouseItem.uuid == item_uuid.lower())
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Elemento di magazzino non trovato")
    return _item_detail_read(item, db)


@router.patch(
    "/items/{item_uuid}",
    response_model=warehouse_schemas.WarehouseItemDetailRead,
)
def update_warehouse_item(
    item_uuid: str,
    item_in: warehouse_schemas.WarehouseItemUpdate,
    db: Session = Depends(get_db),
):
    item = db.scalar(
        select(WarehouseItem)
        .options(joinedload(WarehouseItem.material))
        .where(WarehouseItem.uuid == item_uuid.lower())
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Elemento di magazzino non trovato")
    data = item_in.model_dump(exclude_unset=True)
    for field, value in data.items():
        if isinstance(value, str):
            value = value.strip() or None
        if field in {"peso_u_kg", "peso_1_pz"} and value is not None:
            value = Decimal(str(value))
        setattr(item, field, value)
    if "reserved_for_commessa" in data:
        if item.reserved_for_commessa and item.status == "AVAILABLE":
            item.status = "RESERVED"
        elif not item.reserved_for_commessa and item.status == "RESERVED":
            item.status = "AVAILABLE"
    item.updated_at = datetime.utcnow()
    db.add(item)
    db.commit()
    db.refresh(item)
    return _item_detail_read(item, db)


@router.delete("/items/{item_uuid}", status_code=204)
def delete_warehouse_item(item_uuid: str, db: Session = Depends(get_db)):
    item = db.scalar(select(WarehouseItem).where(WarehouseItem.uuid == item_uuid.lower()))
    if item is None:
        raise HTTPException(status_code=404, detail="Elemento di magazzino non trovato")
    db.delete(item)
    db.commit()
    return Response(status_code=204)


@router.post("/items/bulk-delete", response_model=warehouse_schemas.WarehouseItemBulkDeleteResult)
def delete_warehouse_items(
    request: warehouse_schemas.WarehouseItemBulkRequest,
    db: Session = Depends(get_db),
):
    requested = [str(uuid).strip().lower() for uuid in request.uuids if str(uuid).strip()]
    requested = list(dict.fromkeys(requested))
    if not requested:
        raise HTTPException(status_code=422, detail="Nessun elemento selezionato")
    items = db.scalars(select(WarehouseItem).where(WarehouseItem.uuid.in_(requested))).all()
    found = {item.uuid for item in items}
    for item in items:
        db.delete(item)
    db.commit()
    return {
        "deleted": len(items),
        "requested": len(requested),
        "missing": [uuid for uuid in requested if uuid not in found],
    }


def _warehouse_qr_payload(item: WarehouseItem, material: Material, qr_encoding: str) -> str:
    if qr_encoding == "uuid":
        return item.uuid
    if qr_encoding == "json":
        return json.dumps(
            {
                "domain": "warehouse",
                "uuid": item.uuid,
                "material_code": material.code,
                "ordinal": item.ordinal,
                "uso_materiale": item.uso_materiale or material.uso_materiale,
                "posizione": item.posizione or material.posizione,
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
    return f"https://stqc.mcc.eu/p/{item.uuid}"


@router.get("/items/{item_uuid}/qr.png")
def warehouse_item_qr(
    item_uuid: str,
    qr_encoding: Literal["link", "uuid", "json"] = Query("link", alias="encoding"),
    db: Session = Depends(get_db),
):
    item = db.scalar(
        select(WarehouseItem)
        .options(joinedload(WarehouseItem.material))
        .where(WarehouseItem.uuid == item_uuid.lower())
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Elemento di magazzino non trovato")
    payload = _warehouse_qr_payload(item, item.material, qr_encoding)
    png = base64.b64decode(generate_qr_for_payload(payload))
    return Response(content=png, media_type="image/png", headers={"Cache-Control": "public, max-age=31536000"})


def _extract_uuid(payload: str) -> str:
    match = re.search(r"/p/([0-9a-f-]{36})", payload, re.IGNORECASE)
    if match:
        return match.group(1).lower()
    try:
        data = json.loads(payload)
        value = data.get("uuid") or data.get("id")
        if value:
            return str(value).lower()
    except (json.JSONDecodeError, TypeError):
        pass
    value = payload.strip().lower()
    if re.fullmatch(r"[0-9a-f-]{36}", value, re.IGNORECASE):
        return value
    raise ValueError("QR non riconosciuto")


@router.post("/items/scan", response_model=warehouse_schemas.WarehouseScanResult)
def scan_warehouse_item(
    request: warehouse_schemas.WarehouseScanRequest,
    db: Session = Depends(get_db),
):
    try:
        item_uuid = _extract_uuid(request.payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    item = db.scalar(select(WarehouseItem).where(WarehouseItem.uuid == item_uuid))
    if item is None:
        raise HTTPException(status_code=404, detail="QR non associato a un elemento di magazzino")
    material_row = next(
        (row for row in crud.get_magazzino_list(db=db, limit=10000) if row["material_id"] == item.material_id),
        None,
    )
    if material_row is None:
        raise HTTPException(status_code=404, detail="Materiale collegato non trovato")
    return {"item": _physical_item_read(item), "material": material_row}


def _pdf_text(value: object) -> str:
    return str(value or "-").encode("latin-1", "replace").decode("latin-1")


def _warehouse_labels_pdf(
    material: Material,
    items: list[WarehouseItem],
    label_format: str = "a6",
    custom_text: str | None = None,
    qr_encoding: str = "link",
    qr_size_mm: float | None = None,
) -> bytes:
    formats = {
        "a6": ("P", (105, 148)),
        "rect": ("L", (50, 100)),
        "compact": ("L", (35, 70)),
        "badge": ("L", (60, 90)),
    }
    if label_format not in formats:
        raise ValueError("Formato stampa QR non valido")
    orientation, page_format = formats[label_format]
    pdf = FPDF(orientation=orientation, unit="mm", format=page_format)
    pdf.set_margins(5, 5, 5)
    pdf.set_auto_page_break(False)
    for item in items:
        item_material = item.material or material
        pdf.add_page()
        qr_payload = _warehouse_qr_payload(item, item_material, qr_encoding)
        qr_bytes = base64.b64decode(generate_qr_for_payload(qr_payload))
        identifier = _pdf_text((custom_text or "").strip() or f"{item_material.code} - #{item.ordinal:04d}")

        if label_format == "a6":
            qr_size = qr_size_mm or 78
            qr_x = (105 - qr_size) / 2
            pdf.image(io.BytesIO(qr_bytes), x=qr_x, y=10, w=qr_size, h=qr_size)
            pdf.set_y(8 + qr_size + 5)
            pdf.set_font("Helvetica", "B", 13)
            pdf.cell(0, 8, identifier, align="C", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_y(-12)
            pdf.set_font("Helvetica", "", 6)
            pdf.cell(0, 4, item.uuid, align="C")
        else:
            is_compact = label_format == "compact"
            is_badge = label_format == "badge"
            qr_size = qr_size_mm or (29 if is_compact else 46 if is_badge else 44)
            qr_x = 4
            qr_y = 3 if is_compact else 3
            text_x = qr_x + qr_size + 4
            pdf.image(io.BytesIO(qr_bytes), x=qr_x, y=qr_y, w=qr_size, h=qr_size)
            pdf.set_xy(text_x, qr_y)
            pdf.set_font("Helvetica", "B", 8 if is_compact else 11)
            pdf.multi_cell(0, 4 if is_compact else 5, identifier)
            pdf.set_xy(text_x, qr_y + (14 if is_compact else 20))
            pdf.set_font("Helvetica", "", 5 if is_compact else 6)
            pdf.multi_cell(0, 3, item.uuid)

    return bytes(pdf.output())


@router.get("/materials/{material_id}/labels.pdf")
def warehouse_item_labels(
    material_id: int,
    label_format: Literal["a6", "rect", "compact", "badge"] = Query("a6", alias="format"),
    text: str | None = Query(None, max_length=80),
    qr_encoding: Literal["link", "uuid", "json"] = Query("link", alias="encoding"),
    qr_size_mm: float | None = Query(None, ge=18, le=90, alias="qr_size"),
    db: Session = Depends(get_db),
):
    material = db.get(Material, material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Materiale non trovato")
    items = db.scalars(
        select(WarehouseItem)
        .where(WarehouseItem.material_id == material_id, WarehouseItem.status == "AVAILABLE")
        .order_by(WarehouseItem.ordinal)
    ).all()
    if not items:
        raise HTTPException(status_code=404, detail="Nessun elemento fisico disponibile")
    return Response(
        content=_warehouse_labels_pdf(
            material,
            items,
            label_format,
            text,
            qr_encoding,
            qr_size_mm,
        ),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="qr_{material.code}.pdf"'},
    )


@router.get("/items/{item_uuid}/label.pdf")
def warehouse_single_item_label(
    item_uuid: str,
    label_format: Literal["a6", "rect", "compact", "badge"] = Query("a6", alias="format"),
    text: str | None = Query(None, max_length=80),
    qr_encoding: Literal["link", "uuid", "json"] = Query("link", alias="encoding"),
    qr_size_mm: float | None = Query(None, ge=18, le=90, alias="qr_size"),
    db: Session = Depends(get_db),
):
    item = db.scalar(select(WarehouseItem).where(WarehouseItem.uuid == item_uuid.lower()))
    if item is None:
        raise HTTPException(status_code=404, detail="Elemento di magazzino non trovato")
    return Response(
        content=_warehouse_labels_pdf(
            item.material,
            [item],
            label_format,
            text,
            qr_encoding,
            qr_size_mm,
        ),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="qr_{item.material.code}_{item.ordinal:04d}.pdf"'},
    )


@router.post("/items/labels.pdf")
def warehouse_selected_item_labels(
    request: warehouse_schemas.WarehouseLabelPrintRequest,
    db: Session = Depends(get_db),
):
    requested = [str(uuid).strip().lower() for uuid in request.uuids if str(uuid).strip()]
    requested = list(dict.fromkeys(requested))
    if not requested:
        raise HTTPException(status_code=422, detail="Nessun elemento selezionato")
    items = db.scalars(
        select(WarehouseItem)
        .options(joinedload(WarehouseItem.material))
        .where(WarehouseItem.uuid.in_(requested))
        .order_by(WarehouseItem.material_id, WarehouseItem.ordinal)
    ).all()
    if not items:
        raise HTTPException(status_code=404, detail="Nessun elemento fisico trovato")
    return Response(
        content=_warehouse_labels_pdf(
            items[0].material,
            items,
            request.label_format,
            request.text,
            request.qr_encoding,
            request.qr_size_mm,
        ),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="qr_selezione_magazzino.pdf"'},
    )


@router.delete("/materials/{material_id}", status_code=204)
def delete_material(material_id: int, db: Session = Depends(get_db)):
    """Rimozione fisica di un materiale dal solo dominio magazzino."""
    material = db.get(Material, material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Materiale non trovato")

    # Certificati (tramite receipts di questo materiale)
    receipt_ids = [r.id for r in db.query(Receipt.id).filter(Receipt.material_id == material_id)]
    if receipt_ids:
        db.query(Certificate).filter(Certificate.receipt_id.in_(receipt_ids)).delete(synchronize_session=False)

    db.query(WarehouseCustomValue).filter(WarehouseCustomValue.material_id == material_id).delete(synchronize_session=False)
    db.query(WarehouseItem).filter(WarehouseItem.material_id == material_id).delete(synchronize_session=False)
    db.query(StockMovement).filter(StockMovement.material_id == material_id).delete(synchronize_session=False)
    db.query(Receipt).filter(Receipt.material_id == material_id).delete(synchronize_session=False)
    db.query(Batch).filter(Batch.material_id == material_id).delete(synchronize_session=False)

    db.delete(material)
    db.commit()
